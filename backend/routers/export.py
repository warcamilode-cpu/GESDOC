"""
routers/export.py — Exportar comentarios a Excel o CSV
"""

import io
import csv
from typing import Optional
from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse

import database as db

router = APIRouter()

COLS = ["id", "doc_name", "author", "category", "priority", "status",
        "location_info", "highlighted_text", "content", "created_at"]


def _filas(doc_id: Optional[int]) -> list:
    rows = db.get_all_comments_for_export(doc_id)
    return [
        [str(r.get(c, "") or "") for c in COLS]
        for r in rows
    ]


@router.get("/csv")
def exportar_csv(doc_id: Optional[int] = Query(None)):
    filas = _filas(doc_id)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(COLS)
    w.writerows(filas)
    buf.seek(0)
    nombre = f"comentarios_{doc_id or 'todos'}.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )


@router.get("/excel")
def exportar_excel(doc_id: Optional[int] = Query(None)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    filas = _filas(doc_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comentarios"

    header_fill = PatternFill("solid", fgColor="313244")
    header_font = Font(bold=True, color="89b4fa")

    for col_idx, col_name in enumerate(COLS, 1):
        celda = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
        celda.font = header_font
        celda.fill = header_fill
        celda.alignment = Alignment(horizontal="center")

    prioridad_colores = {"Alta": "f38ba8", "Media": "f9e2af", "Baja": "a6e3a1"}
    estado_colores    = {"Abierto": "89b4fa", "Resuelto": "a6e3a1", "Pendiente": "fab387"}

    for fila_idx, fila in enumerate(filas, 2):
        for col_idx, valor in enumerate(fila, 1):
            celda = ws.cell(row=fila_idx, column=col_idx, value=valor)
            col_name = COLS[col_idx - 1]
            if col_name == "priority" and valor in prioridad_colores:
                celda.fill = PatternFill("solid", fgColor=prioridad_colores[valor])
            elif col_name == "status" and valor in estado_colores:
                celda.fill = PatternFill("solid", fgColor=estado_colores[valor])

    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre = f"comentarios_{doc_id or 'todos'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )
