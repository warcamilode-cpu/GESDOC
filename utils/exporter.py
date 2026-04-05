"""
utils/exporter.py - Export comments to Excel and PDF
"""
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                     Paragraph, Spacer, HRFlowable)
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False


def export_to_excel(comments, output_path, title="Comentarios"):
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl no está instalado. Ejecuta: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comentarios"

    # ── Styles ─────────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill = PatternFill("solid", fgColor="EEF2F8")

    priority_colors = {"Alta": "F38BA8", "Media": "F9E2AF", "Baja": "A6E3A1"}
    status_colors = {"Abierto": "89B4FA", "Resuelto": "A6E3A1", "Pendiente": "FAB387"}

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Title row ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"📋 {title} — Exportado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    title_cell.font = Font(bold=True, size=14, color="1E3A5F")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Header row ─────────────────────────────────────────────────────────
    headers = ["#", "Documento", "Categoría", "Prioridad", "Estado",
               "Texto Resaltado", "Comentario", "Fecha"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[2].height = 22

    # ── Data rows ──────────────────────────────────────────────────────────
    for i, comment in enumerate(comments, 1):
        row = i + 2
        doc_name = comment.get("doc_name", "—")
        values = [
            i,
            doc_name,
            comment.get("category", "—"),
            comment.get("priority", "—"),
            comment.get("status", "—"),
            comment.get("highlighted_text", ""),
            comment.get("content", ""),
            comment.get("created_at", "")[:16] if comment.get("created_at") else "—",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if i % 2 == 0:
                cell.fill = alt_fill

        # Colorize priority & status
        priority = comment.get("priority", "")
        if priority in priority_colors:
            ws.cell(row=row, column=4).fill = PatternFill("solid", fgColor=priority_colors[priority])

        status = comment.get("status", "")
        if status in status_colors:
            ws.cell(row=row, column=5).fill = PatternFill("solid", fgColor=status_colors[status])

        ws.row_dimensions[row].height = 40

    # ── Column widths ──────────────────────────────────────────────────────
    widths = [5, 22, 12, 10, 10, 30, 45, 16]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── Stats sheet ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumen")
    ws2["A1"] = "Resumen de Comentarios"
    ws2["A1"].font = Font(bold=True, size=14, color="1E3A5F")

    from collections import Counter
    priorities = Counter(c.get("priority") for c in comments)
    statuses = Counter(c.get("status") for c in comments)
    categories = Counter(c.get("category") for c in comments)

    ws2["A3"] = "Total Comentarios"
    ws2["B3"] = len(comments)
    ws2["A4"] = "Por Prioridad"
    for i, (k, v) in enumerate(priorities.items(), 5):
        ws2[f"A{i}"] = f"  {k}"
        ws2[f"B{i}"] = v
    r = 5 + len(priorities) + 1
    ws2[f"A{r}"] = "Por Estado"
    for i, (k, v) in enumerate(statuses.items(), r + 1):
        ws2[f"A{i}"] = f"  {k}"
        ws2[f"B{i}"] = v

    for col in ["A", "B"]:
        ws2.column_dimensions[col].width = 25

    wb.save(output_path)
    return output_path


def export_to_pdf(comments, output_path, title="Comentarios"):
    if not HAS_REPORTLAB:
        raise RuntimeError("reportlab no está instalado. Ejecuta: pip install reportlab")

    doc = SimpleDocTemplate(
        output_path, pagesize=A4,
        rightMargin=2 * cm, leftMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm
    )
    styles = getSampleStyleSheet()
    story = []

    # Title
    title_style = ParagraphStyle(
        "Title", parent=styles["Title"],
        fontSize=18, textColor=colors.HexColor("#1e3a5f"),
        spaceAfter=6, alignment=TA_CENTER
    )
    subtitle_style = ParagraphStyle(
        "Subtitle", parent=styles["Normal"],
        fontSize=10, textColor=colors.gray,
        spaceAfter=16, alignment=TA_CENTER
    )
    story.append(Paragraph(f"📋 {title}", title_style))
    story.append(Paragraph(
        f"Exportado el {datetime.now().strftime('%d de %B de %Y a las %H:%M')}",
        subtitle_style
    ))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#89b4fa")))
    story.append(Spacer(1, 12))

    if not comments:
        story.append(Paragraph("No hay comentarios para exportar.", styles["Normal"]))
        doc.build(story)
        return output_path

    # Table headers
    table_data = [["#", "Documento", "Categoría", "Prioridad", "Estado", "Comentario", "Fecha"]]

    cell_style = ParagraphStyle("Cell", parent=styles["Normal"], fontSize=8, leading=10)
    priority_colors_pdf = {
        "Alta": colors.HexColor("#f38ba8"),
        "Media": colors.HexColor("#f9e2af"),
        "Baja": colors.HexColor("#a6e3a1"),
    }

    for i, c in enumerate(comments, 1):
        date_str = (c.get("created_at") or "")[:10]
        table_data.append([
            str(i),
            Paragraph(c.get("doc_name", "—"), cell_style),
            c.get("category", "—"),
            c.get("priority", "—"),
            c.get("status", "—"),
            Paragraph((c.get("content") or "")[:200], cell_style),
            date_str,
        ])

    col_widths = [1.0*cm, 3.5*cm, 2.5*cm, 2.0*cm, 2.0*cm, 5.5*cm, 2.0*cm]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    base_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#eef2f8")]),
        ("PADDING", (0, 0), (-1, -1), 5),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
    ]

    # Color priority column
    for i, c in enumerate(comments, 1):
        pcolor = priority_colors_pdf.get(c.get("priority", ""), None)
        if pcolor:
            base_style.append(("BACKGROUND", (3, i), (3, i), pcolor))

    table.setStyle(TableStyle(base_style))
    story.append(table)

    # Summary
    story.append(Spacer(1, 20))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#cccccc")))
    story.append(Spacer(1, 10))
    from collections import Counter
    priorities = Counter(c.get("priority") for c in comments)
    statuses = Counter(c.get("status") for c in comments)

    summary_style = ParagraphStyle("Summary", parent=styles["Normal"], fontSize=9,
                                    textColor=colors.HexColor("#6c7086"))
    p_text = " | ".join([f"{k}: {v}" for k, v in priorities.items()])
    s_text = " | ".join([f"{k}: {v}" for k, v in statuses.items()])
    story.append(Paragraph(f"Total: {len(comments)} comentarios  ·  Prioridades: {p_text}", summary_style))
    story.append(Paragraph(f"Estados: {s_text}", summary_style))

    doc.build(story)
    return output_path
