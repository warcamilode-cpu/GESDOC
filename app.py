"""
app.py — DocManager Web (Streamlit)

Ejecutar:  streamlit run app.py

Mejoras implementadas:
  1. Subcarpetas ilimitadas en la biblioteca
  2. Comentarios funcionando correctamente (incluye author)
  3. Apartado comentado unificado (PDF y no-PDF) con campo copiar/pegar
  4. Sistema web con Streamlit
  5. Renombrar documentos en la app (renombra también en repositorio/)
  6. Documentos organizados en repositorio/<nombre_carpeta>/
  7. Carpeta renombrada a "repositorio"
"""

import streamlit as st
import os
import base64
import io
import tempfile
import re
from datetime import datetime

import database as db

# ─────────────────────────────────────────────────────────────────────────────
# Configuración de página
# ─────────────────────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="DocManager",
    page_icon="📚",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# CSS global
# ─────────────────────────────────────────────────────────────────────────────

st.markdown("""
<style>
/* ── Base: Times New Roman en toda la interfaz ── */
html, body, [class*="css"] {
    font-family: 'Times New Roman', Times, serif !important;
}
.stApp { background: #1e1e2e; color: #cdd6f4; }
/* Todos los textos de botones, inputs, labels ─── */
button, input, textarea, select, label, p, span, div {
    font-family: 'Times New Roman', Times, serif !important;
}

/* ── Sidebar y comentarios: tamaño 8pt ── */
section[data-testid="stSidebar"],
section[data-testid="stSidebar"] * {
    font-size: 8pt !important;
    font-family: 'Times New Roman', Times, serif !important;
}
.comment-panel, .comment-panel * {
    font-size: 8pt !important;
    font-family: 'Times New Roman', Times, serif !important;
}

/* ── Layout: eliminar márgenes de Streamlit ── */
.block-container { padding-top: 0.8rem !important; padding-bottom: 0.5rem !important; }
header[data-testid="stHeader"] { display: none !important; }

/* ── Sidebar base ── */
section[data-testid="stSidebar"] { background: #181825 !important; }
section[data-testid="stSidebar"] > div:first-child { padding-top: 0.4rem !important; }
section[data-testid="stSidebar"] * { color: #cdd6f4; }

/* ── Compactar TODOS los gaps entre elementos del sidebar ── */
section[data-testid="stSidebar"] [data-testid="stVerticalBlock"] {
    gap: 0px !important;
}
section[data-testid="stSidebar"] [data-testid="element-container"] {
    padding-top: 1px !important;
    padding-bottom: 1px !important;
}
/* Reducir espacio alrededor de botones en sidebar */
section[data-testid="stSidebar"] [data-testid="stHorizontalBlock"] {
    gap: 3px !important;
    margin-bottom: 2px !important;
    margin-top: 0px !important;
}
section[data-testid="stSidebar"] .stTextInput {
    margin-bottom: 3px !important;
}
section[data-testid="stSidebar"] .stTextInput > div {
    margin-bottom: 0 !important;
}
section[data-testid="stSidebar"] .stTextInput > div > div > input {
    padding: 3px 8px !important;
    height: 28px !important;
    font-size: 8pt !important;
}

/* ── Botones ── */
.stButton > button {
    background: #313244;
    color: #cdd6f4;
    border: 1px solid #45475a;
    border-radius: 5px;
    transition: all .15s;
    font-family: 'Times New Roman', Times, serif !important;
}
.stButton > button:hover { background: #45475a; border-color: #89b4fa; color: #89b4fa; }

button[kind="primary"] {
    background: #89b4fa !important; color: #1e1e2e !important;
    border: none !important; font-weight: 600;
}
button[kind="primary"]:hover { background: #74c7ec !important; }

/* ── Inputs globales ── */
.stTextInput > div > div > input,
.stTextArea > div > div > textarea,
.stSelectbox > div > div {
    background: #313244 !important; color: #cdd6f4 !important;
    border: 1px solid #45475a !important; border-radius: 5px !important;
}

/* ── Expander ── */
.streamlit-expanderHeader {
    background: #24273a !important; color: #89b4fa !important;
    border-radius: 5px; font-weight: 600;
}

/* ── Sidebar: búsqueda y "Añadir documento" compactos ── */
section[data-testid="stSidebar"] .stButton > button {
    font-size: 8pt !important;
    padding: 3px 8px !important;
    min-height: 26px !important;
    text-align: left;
}

/* ── Botón toggle de carpeta: primera columna de fila horizontal ── */
section[data-testid="stSidebar"] [data-testid="stHorizontalBlock"]
    > div:first-child .stButton > button {
    background: #1e3a5f !important;
    border: 1px solid #89b4fa33 !important;
    color: #89b4fa !important;
    font-weight: 700 !important;
    text-align: left !important;
    font-size: 8pt !important;
    font-family: 'Courier New', Courier, monospace !important;
    white-space: pre !important;
    border-radius: 5px !important;
    width: 100% !important;
    padding: 4px 6px !important;
    min-height: 26px !important;
    margin-bottom: 1px !important;
    overflow: hidden !important;
    text-overflow: ellipsis !important;
}
section[data-testid="stSidebar"] [data-testid="stHorizontalBlock"]
    > div:first-child .stButton > button:hover {
    background: #24508a !important; border-color: #89b4fa !important;
}

/* ── Botones acción (cols 2,3,4 de la fila) ── */
section[data-testid="stSidebar"] [data-testid="stHorizontalBlock"]
    > div:not(:first-child) .stButton > button {
    background: transparent !important;
    border: 1px solid #45475a44 !important;
    color: #6c7086 !important;
    font-size: 7pt !important;
    padding: 0px 2px !important;
    min-height: 26px !important;
    max-height: 26px !important;
    border-radius: 3px !important;
    margin: 0 !important;
}
section[data-testid="stSidebar"] [data-testid="stHorizontalBlock"]
    > div:not(:first-child) .stButton > button:hover {
    background: #313244 !important; color: #cdd6f4 !important;
    border-color: #89b4fa !important;
}



/* ── Divider ── */
hr { border-color: #313244; }

/* ── Alineación izquierda: todos los botones sidebar ── */
section[data-testid="stSidebar"] button {
    text-align: left !important;
    justify-content: flex-start !important;
}
section[data-testid="stSidebar"] button p,
section[data-testid="stSidebar"] button span {
    text-align: left !important;
    width: 100% !important;
}


/* ── Card ── */
.comment-card {
    background: #24273a;
    border: 1px solid #313244;
    border-radius: 10px;
    padding: 14px;
    margin-bottom: 10px;
}
.comment-chip {
    display: inline-block;
    border-radius: 5px;
    padding: 2px 8px;
    font-size: 11px;
    font-weight: 600;
    margin-right: 4px;
}
.highlighted-box {
    background: #1e2030;
    border-left: 3px solid #89b4fa;
    padding: 8px 12px;
    border-radius: 0 6px 6px 0;
    color: #a6e3a1;
    font-style: italic;
    font-size: 13px;
    margin: 8px 0;
}
.doc-btn {
    display: block;
    width: 100%;
    text-align: left;
    background: #24273a;
    border: 1px solid #313244;
    border-radius: 6px;
    padding: 6px 10px;
    color: #cdd6f4;
    font-size: 13px;
    cursor: pointer;
    margin-bottom: 2px;
    transition: background .15s;
}
.doc-btn:hover { background: #313244; border-color: #89b4fa; color: #89b4fa; }
.doc-btn.active { background: #1e3a5f; border-color: #89b4fa; color: #89b4fa; }

/* ── Scrollbar ── */
::-webkit-scrollbar { width: 6px; }
::-webkit-scrollbar-track { background: #1e1e2e; }
::-webkit-scrollbar-thumb { background: #45475a; border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: #6c7086; }

/* ── Status colors ── */
.status-por_revisar { color: #89b4fa; }
.status-en_progreso  { color: #fab387; }
.status-revisado     { color: #f9e2af; }
.status-aprobado     { color: #a6e3a1; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# Session State
# ─────────────────────────────────────────────────────────────────────────────

_DEFAULTS = {
    "current_doc_id":     None,
    "show_add_comment":   False,
    "show_add_folder":    False,
    "show_add_subfolder": None,
    "rename_doc_id":      None,
    "edit_comment_id":    None,
    "confirm_delete_doc": None,
    "confirm_delete_fold":None,
    "expand_folder":      set(),
    "comments_visible":   True,   # panel comentarios visible/oculto
}

def _init_state():
    for k, v in _DEFAULTS.items():
        if k not in st.session_state:
            st.session_state[k] = v

# ─────────────────────────────────────────────────────────────────────────────
# Constantes de estilo
# ─────────────────────────────────────────────────────────────────────────────

FMT_ICONS = {"pdf":"📕","docx":"📘","doc":"📘","md":"📗","txt":"📄"}
PRI_COLORS = {"Alta":"#f38ba8","Media":"#f9e2af","Baja":"#a6e3a1"}
STS_COLORS = {"Abierto":"#89b4fa","Resuelto":"#a6e3a1","Pendiente":"#fab387"}
CAT_ICONS  = {
    "General":"💬","Importante":"⚠️","Pregunta":"❓",
    "Corrección":"✏️","Referencia":"📎","Tarea":"✅",
}
DOC_STATUS_COLORS = {
    "Por revisar":"#89b4fa","En progreso":"#fab387",
    "Revisado":"#f9e2af","Aprobado":"#a6e3a1",
}

# ─────────────────────────────────────────────────────────────────────────────
# Visor de documentos
# ─────────────────────────────────────────────────────────────────────────────

def _pdf_html(path: str) -> str:
    with open(path, "rb") as f:
        b64 = base64.b64encode(f.read()).decode()
    return (
        '<div style="width:100%;height:calc(100vh - 110px);min-height:700px;">'
        f'<iframe src="data:application/pdf;base64,{b64}" '
        f'width="100%" height="100%" style="border:none;border-radius:8px;" '
        f'type="application/pdf"></iframe></div>'
    )


def _docx_html(path: str) -> str:
    try:
        import mammoth
        with open(path, "rb") as f:
            result = mammoth.convert_to_html(f)
        return (
            f'<div style="font-family:Georgia,serif;font-size:14px;'
            f'line-height:1.7;padding:24px;background:#fafafa;color:#1e1e2e;'
            f'border-radius:8px;height:calc(100vh - 110px);min-height:700px;overflow-y:auto;">'
            f'{result.value}</div>'
        )
    except ImportError:
        return "<p style='color:#f38ba8'>Instala <b>mammoth</b>: pip install mammoth</p>"
    except Exception as e:
        return f"<p style='color:#f38ba8'>Error al leer DOCX: {e}</p>"


def render_viewer(doc: dict):
    path = doc.get("path", "")
    fmt  = doc.get("format", "txt").lower()

    if not os.path.isfile(path):
        st.error(f"⚠️ Archivo no encontrado:\n`{path}`")
        return

    if fmt == "pdf":
        st.markdown(_pdf_html(path), unsafe_allow_html=True)

    elif fmt in ("docx", "doc"):
        st.markdown(_docx_html(path), unsafe_allow_html=True)

    elif fmt == "md":
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            content = f.read()
        # render markdown directly — no wrapper div
        st.markdown(content)

    else:  # txt
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            content = f.read()
        st.text_area(
            "Contenido del archivo",
            value=content,
            height=780,
            disabled=True,
            label_visibility="collapsed",
        )


# ─────────────────────────────────────────────────────────────────────────────
# Panel de comentarios
# ─────────────────────────────────────────────────────────────────────────────

def _chip(text, color):
    return (f'<span style="display:inline-block;border-radius:4px;padding:1px 6px;' 
            f'font-size:8pt;font-weight:600;margin-right:3px;' 
            f'background:{color}22;color:{color};">{text}</span>')


def _comment_card_html(c: dict, is_reply: bool = False) -> str:
    """HTML abreviado de una tarjeta de comentario (preview de 120 chars)."""
    pri   = c.get("priority","Media")
    sts   = c.get("status","Abierto")
    cat   = c.get("category","General")
    p_col = PRI_COLORS.get(pri,"#cdd6f4")
    s_col = STS_COLORS.get(sts,"#cdd6f4")
    cat_i = CAT_ICONS.get(cat,"💬")
    date  = (c.get("created_at") or "")[:10]
    author = c.get("author","") or ""

    chips = (
        _chip(f"{cat_i} {cat}", p_col) +
        _chip(f"● {pri}", p_col) +
        _chip(f"◆ {sts}", s_col) +
        f'<span style="float:right;color:#6c7086;font-size:8pt;">{date}</span>'
    )

    meta = ""
    if author:
        meta += f'<span style="color:#89b4fa;font-size:8pt;margin-right:8px;">👤 {author}</span>'
    if c.get("location_info"):
        meta += f'<span style="color:#a6adc8;font-size:8pt;">📍 {c["location_info"]}</span>'

    # Preview: máx 120 chars
    raw = (c.get("content") or "")
    preview = raw[:120] + ("…" if len(raw) > 120 else "")

    bg   = "#1e2535" if is_reply else "#24273a"
    ml   = "margin-left:16px;" if is_reply else ""
    border = "border-left:3px solid #45475a;" if is_reply else "border:1px solid #313244;"

    return (
        f'<div style="background:{bg};{border}{ml}' 
        f'border-radius:8px;padding:10px 12px;margin-bottom:6px;">' 
        f'<div style="margin-bottom:4px;">{chips}</div>'
        + (f'<div style="margin-bottom:3px;">{meta}</div>' if meta else "")
        + (
            f'<div style="background:#1e2030;border-left:3px solid #89b4fa;' 
            f'padding:5px 8px;border-radius:0 4px 4px 0;' 
            f'color:#a6e3a1;font-style:italic;font-size:8pt;margin-bottom:5px;">' 
            f'📋 {(c.get("highlighted_text","") or "")[:150]}' 
            f'{"…" if len(c.get("highlighted_text","") or "") > 150 else ""}' 
            f'</div>'
            if c.get("highlighted_text") else ""
        )
        + f'<div style="color:#cdd6f4;font-size:8pt;line-height:1.5;">{preview}</div>'
        + f'</div>'
    )


def _comment_full_html(c: dict) -> str:
    """HTML completo de un comentario expandido."""
    pri   = c.get("priority","Media")
    sts   = c.get("status","Abierto")
    cat   = c.get("category","General")
    p_col = PRI_COLORS.get(pri,"#cdd6f4")
    s_col = STS_COLORS.get(sts,"#cdd6f4")
    cat_i = CAT_ICONS.get(cat,"💬")
    date  = (c.get("created_at") or "")[:16]
    author = c.get("author","") or ""

    hl_block = ""
    if c.get("highlighted_text"):
        hl_block = (
            f'<div style="background:#1e2030;border-left:3px solid #89b4fa;' 
            f'padding:8px 10px;border-radius:0 6px 6px 0;' 
            f'color:#a6e3a1;font-style:italic;font-size:8pt;' 
            f'margin:6px 0;">📋 {c["highlighted_text"]}</div>'
        )

    return (
        f'<div style="background:#1a1f2e;border:1px solid #45475a;' 
        f'border-radius:8px;padding:14px;margin-bottom:6px;">' 
        f'<div style="margin-bottom:6px;">' 
        + _chip(f"{cat_i} {cat}", p_col) + _chip(f"● {pri}", p_col) + _chip(f"◆ {sts}", s_col)
        + f'<span style="float:right;color:#6c7086;font-size:8pt;">{date}</span></div>'
        + (f'<div style="color:#89b4fa;font-size:8pt;margin-bottom:2px;">👤 {author}</div>' if author else "")
        + (f'<div style="color:#a6adc8;font-size:8pt;margin-bottom:4px;">📍 {c["location_info"]}</div>' if c.get("location_info") else "")
        + hl_block
        + f'<div style="color:#cdd6f4;font-size:8pt;line-height:1.6;white-space:pre-wrap;">{c.get("content","")}</div>'
        + f'</div>'
    )


def render_comments(doc: dict):
    doc_id = doc["id"]

    # ── Encabezado ────────────────────────────────────────────────────────────
    all_c      = db.get_comments(doc_id)
    open_count = sum(1 for c in all_c if c.get("status") == "Abierto")

    # comment-panel: no div wrapper (causes React errors)
    c1, c2 = st.columns([3, 1])
    c1.markdown('<span style="font-size:10pt;font-weight:700;color:#89b4fa;">💬 Comentarios</span>',
                unsafe_allow_html=True)
    c1.markdown(f'<span style="font-size:8pt;color:#6c7086;">{len(all_c)} total · {open_count} abiertos</span>',
                unsafe_allow_html=True)
    if c2.button("➕ Nuevo", key="btn_add_cmt", use_container_width=True, type="primary"):
        st.session_state.show_add_comment = not st.session_state.show_add_comment
        st.session_state.edit_comment_id  = None
        st.rerun()

    if st.session_state.show_add_comment:
        _add_comment_form(doc_id)

    # ── Filtros ───────────────────────────────────────────────────────────────
    with st.expander("🔍 Filtros", expanded=False):
        fc1, fc2, fc3 = st.columns(3)
        cat_f = fc1.selectbox("Cat",  ["Todas"] + db.COMMENT_CATEGORIES, key="cf_cat", label_visibility="collapsed")
        pri_f = fc2.selectbox("Pri",  ["Todas"] + db.COMMENT_PRIORITIES, key="cf_pri", label_visibility="collapsed")
        sts_f = fc3.selectbox("Sts",  ["Todos"]  + db.COMMENT_STATUSES,  key="cf_sts", label_visibility="collapsed")
        srch  = st.text_input("Buscar", key="cf_srch", placeholder="🔍 Buscar...")

    root_comments = db.get_root_comments(
        doc_id,
        category=cat_f if cat_f != "Todas" else None,
        priority=pri_f if pri_f != "Todas" else None,
        status  =sts_f if sts_f != "Todos"  else None,
        search  =srch or None,
    )

    if not root_comments:
        st.info("No hay comentarios que coincidan.")
        return

    for c in root_comments:
        _render_comment_thread(c, doc_id, is_reply=False)


def _add_comment_form(doc_id: int):
    with st.form("form_add_comment", clear_on_submit=True):
        st.markdown('<span style="color:#89b4fa;font-size:8pt;font-weight:700;">➕ Nuevo comentario</span>',
                    unsafe_allow_html=True)
        ca1, ca2 = st.columns(2)
        author = ca1.text_input("Autor", placeholder="Tu nombre...", label_visibility="collapsed",
                                key="nc_author")
        loc    = ca2.text_input("Página/Ubicación", placeholder="Ej: Página 5", label_visibility="collapsed",
                                key="nc_loc")
        cc1, cc2, cc3 = st.columns(3)
        cat = cc1.selectbox("Categoría", db.COMMENT_CATEGORIES, key="nc_cat",
                            label_visibility="collapsed")
        pri = cc2.selectbox("Prioridad",  db.COMMENT_PRIORITIES, key="nc_pri",
                            label_visibility="collapsed")
        sts = cc3.selectbox("Estado",     db.COMMENT_STATUSES,   key="nc_sts",
                            label_visibility="collapsed")
        hl  = st.text_area("Apartado comentado", height=60, key="nc_hl",
                           placeholder="Texto del documento que se comenta...",
                           label_visibility="collapsed")
        body = st.text_area("Comentario *", height=80, key="nc_body",
                            placeholder="Escribe tu comentario aquí...",
                            label_visibility="collapsed")
        sb1, sb2 = st.columns(2)
        if sb1.form_submit_button("💾 Guardar", use_container_width=True, type="primary"):
            if body.strip():
                db.add_comment(
                    doc_id           = doc_id,
                    content          = body.strip(),
                    category         = cat,
                    priority         = pri,
                    status           = sts,
                    location_info    = loc.strip(),
                    highlighted_text = hl.strip(),
                    author           = author.strip(),
                )
                st.session_state.show_add_comment = False
                st.rerun()
        if sb2.form_submit_button("Cancelar", use_container_width=True):
            st.session_state.show_add_comment = False
            st.rerun()


def _edit_comment_form(c: dict):
    with st.form(f"form_edit_comment_{c['id']}", clear_on_submit=False):
        st.markdown('<span style="color:#89b4fa;font-size:8pt;font-weight:700;">✏️ Editar comentario</span>',
                    unsafe_allow_html=True)
        ea1, ea2 = st.columns(2)
        ea  = ea1.text_input("Autor",    value=c.get("author",""),        label_visibility="collapsed")
        loc = ea2.text_input("Ubicación",value=c.get("location_info",""), label_visibility="collapsed")
        ec1, ec2, ec3 = st.columns(3)
        ecat = ec1.selectbox("Categoría", db.COMMENT_CATEGORIES,
                             index=_safe_idx(db.COMMENT_CATEGORIES, c.get("category","General")),
                             key=f"ec_cat_{c['id']}", label_visibility="collapsed")
        epri = ec2.selectbox("Prioridad",  db.COMMENT_PRIORITIES,
                             index=_safe_idx(db.COMMENT_PRIORITIES, c.get("priority","Media")),
                             key=f"ec_pri_{c['id']}", label_visibility="collapsed")
        ests = ec3.selectbox("Estado",     db.COMMENT_STATUSES,
                             index=_safe_idx(db.COMMENT_STATUSES, c.get("status","Abierto")),
                             key=f"ec_sts_{c['id']}", label_visibility="collapsed")
        ehl   = st.text_area("Apartado comentado", value=c.get("highlighted_text",""),
                             height=60, label_visibility="collapsed")
        econt = st.text_area("Comentario", value=c.get("content",""),
                             height=80, label_visibility="collapsed")
        es1, es2 = st.columns(2)
        if es1.form_submit_button("💾 Guardar", use_container_width=True, type="primary"):
            if econt.strip():
                db.update_comment(
                    c["id"],
                    content          = econt.strip(),
                    category         = ecat,
                    priority         = epri,
                    status           = ests,
                    location_info    = loc,
                    author           = ea.strip(),
                    highlighted_text = ehl.strip(),
                )
                st.session_state.edit_comment_id = None
                st.rerun()
        if es2.form_submit_button("Cancelar", use_container_width=True):
            st.session_state.edit_comment_id = None
            st.rerun()


def _render_comment_thread(c: dict, doc_id: int, is_reply: bool = False):
    """Renderiza un comentario con sus respuestas en hilo colapsable."""
    cid         = c["id"]
    expanded_k  = f"cmt_expanded_{cid}"
    reply_form_k= f"reply_form_{cid}"
    replies     = db.get_replies(cid)
    n_replies   = len(replies)

    # ── Tarjeta abreviada + botón expandir ────────────────────────────────────
    st.markdown(_comment_card_html(c, is_reply=is_reply), unsafe_allow_html=True)

    # Fila de acciones compacta
    btn_cols = st.columns([2, 2, 2, 2, 2, 1])
    is_open  = st.session_state.get(expanded_k, False)

    # Botón expandir/colapsar (muestra cuenta de respuestas)
    reply_badge = f" ({n_replies})" if n_replies else ""
    expand_lbl  = f"▾ Ver{reply_badge}" if is_open else f"▸ Ver{reply_badge}"
    if btn_cols[0].button(expand_lbl, key=f"exp_{cid}", use_container_width=True):
        st.session_state[expanded_k] = not is_open
        st.rerun()

    if btn_cols[1].button("💬 Resp.", key=f"rep_{cid}", use_container_width=True,
                           help="Responder en hilo"):
        st.session_state[reply_form_k] = not st.session_state.get(reply_form_k, False)
        st.rerun()

    new_sts = "Abierto" if c.get("status") == "Resuelto" else "Resuelto"
    btn_lbl = "🔄" if c.get("status") == "Resuelto" else "✅"
    if btn_cols[2].button(btn_lbl, key=f"res_{cid}", use_container_width=True,
                           help="Resolver / Reabrir"):
        db.update_comment(cid, status=new_sts)
        st.rerun()

    if btn_cols[3].button("✏️", key=f"edit_{cid}", use_container_width=True, help="Editar"):
        if st.session_state.edit_comment_id == cid:
            st.session_state.edit_comment_id = None
        else:
            st.session_state.edit_comment_id  = cid
            st.session_state.show_add_comment = False
        st.rerun()

    if btn_cols[4].button("📋", key=f"copy_{cid}", use_container_width=True, help="Copiar texto"):
        st.toast("Copia el texto con Ctrl+C.")

    if btn_cols[5].button("🗑", key=f"del_{cid}", use_container_width=True, help="Eliminar"):
        db.delete_comment(cid)
        st.rerun()

    # ── Contenido expandido ───────────────────────────────────────────────────
    if is_open:
        st.markdown(_comment_full_html(c), unsafe_allow_html=True)

    # ── Formulario de edición ─────────────────────────────────────────────────
    if st.session_state.edit_comment_id == cid:
        _edit_comment_form(c)

    # ── Formulario de respuesta ───────────────────────────────────────────────
    if st.session_state.get(reply_form_k, False):
        _reply_form(doc_id, parent_id=cid, form_key=reply_form_k)

    # ── Respuestas en hilo (indentadas) ───────────────────────────────────────
    if n_replies > 0:
        for reply in replies:
            _render_comment_thread(reply, doc_id, is_reply=True)

    st.markdown('<hr style="border-color:#313244;margin:4px 0;">', unsafe_allow_html=True)


def _reply_form(doc_id: int, parent_id: int, form_key: str):
    """Formulario inline para responder a un comentario."""
    with st.form(f"form_reply_{parent_id}", clear_on_submit=True):
        st.markdown('<span style="color:#cba6f7;font-size:8pt;font-weight:700;">💬 Responder en hilo</span>',
                    unsafe_allow_html=True)
        r1, r2 = st.columns(2)
        author  = r1.text_input("Autor", placeholder="Tu nombre", label_visibility="collapsed",
                                 key=f"reply_author_{parent_id}")
        content_txt = st.text_area("Respuesta", height=70, placeholder="Escribe tu respuesta...",
                                    label_visibility="collapsed", key=f"reply_content_{parent_id}")
        sb1, sb2 = st.columns(2)
        if sb1.form_submit_button("💾 Guardar", use_container_width=True, type="primary"):
            if content_txt.strip():
                db.add_comment(
                    doc_id    = doc_id,
                    content   = content_txt.strip(),
                    author    = author.strip(),
                    parent_id = parent_id,
                )
                st.session_state[form_key] = False
                st.rerun()
        if sb2.form_submit_button("Cancelar", use_container_width=True):
            st.session_state[form_key] = False
            st.rerun()


def render_sidebar():
    st.sidebar.markdown(
        '<div style="padding:8px 4px 5px 4px;margin-bottom:4px;">'
        '<span style="font-size:11pt;font-weight:700;color:#89b4fa;'
        'letter-spacing:.5px;text-transform:uppercase;'
        'font-family:Times New Roman,serif;">'
        '🗂 ADMINISTRADOR DE DOCUMENTOS</span></div>',
        unsafe_allow_html=True
    )

    # ── Buscar ──
    search_q = st.sidebar.text_input(
        "🔍", placeholder="Buscar documentos...", key="sb_search", label_visibility="collapsed"
    )

    if search_q:
        _render_search_results(search_q)
        return

    # ── Botón: Añadir documento ──
    if st.sidebar.button("➕ Añadir documento", key="sb_add_doc", use_container_width=True):
        st.session_state.show_add_doc_form = not st.session_state.get("show_add_doc_form", False)

    if st.session_state.get("show_add_doc_form", False):
        _add_document_form()

    st.sidebar.markdown('<div style="height:1px;background:#313244;margin:2px 0 3px 0;"></div>', unsafe_allow_html=True)
    st.sidebar.markdown('<span style="font-size:8pt;font-weight:700;color:#6c7086;letter-spacing:1px;text-transform:uppercase;font-family:Times New Roman,serif;">📂 BIBLIOTECA</span>', unsafe_allow_html=True)

    # ── Botón: Nueva carpeta raíz ──
    if st.sidebar.button("📁 Nueva carpeta", key="sb_new_folder", use_container_width=True):
        st.session_state.show_add_folder = not st.session_state.show_add_folder

    if st.session_state.show_add_folder:
        _new_folder_form(parent_id=None)

    # ── Árbol de carpetas ──
    all_folders = db.get_all_folders()
    _render_folder_tree(all_folders, parent_id=None, depth=0)

    # ── Sin carpeta ──
    unfoldered = db.get_unfoldered_documents()
    if unfoldered:
        with st.sidebar.expander(f"📄 Sin carpeta ({len(unfoldered)})", expanded=False):
            for doc in unfoldered:
                _render_doc_item(doc, all_folders)


# ─── Árbol de carpetas colapsable ───────────────────────────────────────────
# Colores por nivel de profundidad
_DEPTH_COLORS = ["#89b4fa", "#cba6f7", "#f38ba8", "#fab387", "#f9e2af"]
_DEPTH_BG     = ["#1e3a5f", "#2a1f3d", "#3d1a1f", "#3d2a1a", "#3d3a1a"]


def _folder_expanded_key(folder_id):
    return f"tree_expanded_{folder_id}"


def _render_folder_tree(all_folders, parent_id, depth, ancestor_has_more=None):
    """
    Árbol colapsable puro — sin HTML suelto.
    Los conectores de árbol van embebidos en el label del botón.
    ancestor_has_more: lista de booleans indicando si cada ancestro tiene más hermanos.
    """
    if ancestor_has_more is None:
        ancestor_has_more = []

    children = [f for f in all_folders if f.get("parent_id") == parent_id]
    if not children:
        return

    for i, folder in enumerate(children):
        fid          = folder["id"]
        docs         = db.get_documents_in_folder(fid)
        sub_children = [f for f in all_folders if f.get("parent_id") == fid]
        doc_cnt      = len(docs)
        expanded_key = _folder_expanded_key(fid)
        is_last      = (i == len(children) - 1)

        if expanded_key not in st.session_state:
            st.session_state[expanded_key] = False

        is_open = st.session_state[expanded_key]
        arrow   = "▾" if is_open else "▸"
        color   = _DEPTH_COLORS[min(depth, len(_DEPTH_COLORS) - 1)]

        # ── Construir prefijo del árbol ──────────────────────────────────────
        if depth == 0:
            prefix = ""
        else:
            # Líneas verticales de los ancestros
            trunk = ""
            for has_more in ancestor_has_more:
                trunk += "│  " if has_more else "   "
            branch = "└─ " if is_last else "├─ "
            prefix = trunk + branch

        toggle_label = f"{prefix}{arrow} {folder['icon']} {folder['name']}  ({doc_cnt})"

        # ── Fila: [título] [✏] [+] [🗑] ─────────────────────────────────────
        col_title, col_e, col_p, col_d = st.sidebar.columns([7, 1, 1, 1])

        if col_title.button(
            toggle_label,
            key=f"toggle_{fid}",
            use_container_width=True,
            help="Expandir / Colapsar",
        ):
            st.session_state[expanded_key] = not is_open
            st.rerun()

        if col_e.button("✏", key=f"rf_{fid}", help="Renombrar", use_container_width=True):
            st.session_state[f"rename_folder_{fid}"] = \
                not st.session_state.get(f"rename_folder_{fid}", False)
        if col_p.button("+", key=f"sf_{fid}", help="Nueva subcarpeta", use_container_width=True):
            st.session_state.show_add_subfolder = \
                fid if st.session_state.show_add_subfolder != fid else None
        if col_d.button("🗑", key=f"df_{fid}", help="Eliminar", use_container_width=True):
            st.session_state[f"confirm_del_f_{fid}"] = True

        # ── Contenido (solo si expandido) ────────────────────────────────────
        if is_open:
            # Confirmación borrar
            if st.session_state.get(f"confirm_del_f_{fid}", False):
                st.sidebar.warning(f"¿Eliminar '{folder['name']}'?")
                cc1, cc2 = st.sidebar.columns(2)
                if cc1.button("Sí", key=f"conf_del_f_{fid}"):
                    db.delete_folder(fid)
                    st.rerun()
                if cc2.button("No", key=f"canc_del_f_{fid}"):
                    st.session_state[f"confirm_del_f_{fid}"] = False
                    st.rerun()

            # Renombrar carpeta
            if st.session_state.get(f"rename_folder_{fid}", False):
                with st.sidebar.form(f"form_rename_f_{fid}"):
                    icons_opts = ["📁","📂","🗂","📚","📋","🗃","📌","⭐","💼"]
                    new_fn  = st.text_input("Nuevo nombre", value=folder["name"])
                    new_ico = st.selectbox("Icono", icons_opts,
                                           index=icons_opts.index(folder["icon"])
                                                 if folder["icon"] in icons_opts else 0)
                    rf1, rf2 = st.columns(2)
                    if rf1.form_submit_button("💾", use_container_width=True):
                        db.update_folder(fid, name=new_fn.strip(), icon=new_ico)
                        st.session_state[f"rename_folder_{fid}"] = False
                        st.rerun()
                    if rf2.form_submit_button("✕", use_container_width=True):
                        st.session_state[f"rename_folder_{fid}"] = False
                        st.rerun()

            # Nueva subcarpeta
            if st.session_state.show_add_subfolder == fid:
                _new_folder_form(parent_id=fid, key_suffix=str(fid))

            # Documentos de esta carpeta
            all_f = db.get_all_folders()
            for doc in docs:
                _render_doc_item(doc, all_f)

            # Subcarpetas recursivas — pasamos si este nivel tiene más hermanos
            _render_folder_tree(
                all_folders,
                parent_id=fid,
                depth=depth + 1,
                ancestor_has_more=ancestor_has_more + [not is_last],
            )


def _new_folder_form(parent_id=None, key_suffix="root"):
    with st.form(f"form_new_folder_{key_suffix}", clear_on_submit=True):
        icons_list = ["📁","📂","🗂","📚","📋","🗃","📌","⭐","💼","🔬","📐","🧩"]
        col1, col2 = st.columns([3, 1])
        fname = col1.text_input("Nombre", placeholder="Nombre de la carpeta...", label_visibility="collapsed")
        ficon = col2.selectbox("Icono", icons_list, label_visibility="collapsed")
        sf1, sf2 = st.columns(2)
        if sf1.form_submit_button("➕ Crear", use_container_width=True, type="primary"):
            if fname.strip():
                db.add_folder(fname.strip(), icon=ficon, parent_id=parent_id)
                st.session_state.show_add_folder   = False
                st.session_state.show_add_subfolder = None
                st.rerun()
        if sf2.form_submit_button("Cancelar", use_container_width=True):
            st.session_state.show_add_folder    = False
            st.session_state.show_add_subfolder = None
            st.rerun()


def _add_document_form():
    with st.sidebar.form("form_add_doc", clear_on_submit=True):
        uploaded = st.file_uploader(
            "Seleccionar archivo",
            type=["pdf","docx","doc","md","txt"],
            key="file_uploader",
        )
        doc_name_input = st.text_input("Nombre (opcional)", placeholder="Nombre del documento...")

        all_folders    = db.get_all_folders()
        folder_labels  = ["Sin carpeta"] + [
            f"{'  ' * _folder_depth(all_folders, f['id'])}{f['icon']} {f['name']}"
            for f in all_folders
        ]
        folder_choice = st.selectbox("Asignar a carpeta", folder_labels)

        s1, s2 = st.columns(2)
        submitted = s1.form_submit_button("💾 Guardar", use_container_width=True, type="primary")
        cancelled = s2.form_submit_button("✕",          use_container_width=True)

        if submitted and uploaded:
            ext  = os.path.splitext(uploaded.name)[1].lower().lstrip(".")
            name = doc_name_input.strip() or os.path.splitext(uploaded.name)[0]

            folder_id   = None
            folder_name = None
            if folder_choice != "Sin carpeta":
                idx         = folder_labels.index(folder_choice) - 1
                folder_id   = all_folders[idx]["id"]
                folder_name = all_folders[idx]["name"]

            with tempfile.NamedTemporaryFile(delete=False, suffix=f".{ext}") as tmp:
                tmp.write(uploaded.read())
                tmp_path = tmp.name

            doc_id = db.add_document(name, tmp_path, ext, folder_name=folder_name)
            try:
                os.unlink(tmp_path)
            except Exception:
                pass

            if doc_id and folder_id:
                db.assign_document_to_folder(doc_id, folder_id)

            st.session_state.show_add_doc_form = False
            st.success(f"✅ '{name}' añadido")
            st.rerun()

        if cancelled:
            st.session_state.show_add_doc_form = False
            st.rerun()


def _render_doc_item(doc: dict, all_folders: list):
    doc_id  = doc["id"]
    fmt     = doc.get("format","txt").lower()
    icon    = FMT_ICONS.get(fmt,"📄")
    is_cur  = st.session_state.current_doc_id == doc_id
    name    = doc.get("name","Documento")
    short   = name[:28] + "…" if len(name) > 28 else name

    sts_col = DOC_STATUS_COLORS.get(doc.get("status","Por revisar"),"#89b4fa")

    dc1, dc2 = st.columns([5, 1])

    btn_style = "primary" if is_cur else "secondary"
    if dc1.button(
        f"{icon} {short}",
        key        = f"open_doc_{doc_id}",
        use_container_width = True,
        help       = name,
    ):
        st.session_state.current_doc_id   = doc_id
        st.session_state.show_add_comment = False
        st.session_state.edit_comment_id  = None
        db.update_last_opened(doc_id)
        st.rerun()

    if dc2.button("⋮", key=f"menu_{doc_id}", help="Opciones"):
        st.session_state[f"doc_menu_{doc_id}"] = \
            not st.session_state.get(f"doc_menu_{doc_id}", False)

    if st.session_state.get(f"doc_menu_{doc_id}", False):
        _doc_context_menu(doc, all_folders)


def _doc_context_menu(doc: dict, all_folders: list):
    doc_id = doc["id"]

    with st.container():
        # ── Renombrar ──
        if st.button("✏️ Renombrar", key=f"rename_btn_{doc_id}", use_container_width=True):
            st.session_state[f"renaming_{doc_id}"] = True
            st.session_state[f"doc_menu_{doc_id}"] = False

        if st.session_state.get(f"renaming_{doc_id}", False):
            with st.form(f"form_rename_doc_{doc_id}", clear_on_submit=True):
                new_name = st.text_input("Nuevo nombre", value=doc.get("name",""))
                rn1, rn2 = st.columns(2)
                if rn1.form_submit_button("💾", use_container_width=True):
                    if new_name.strip():
                        db.rename_document(doc_id, new_name.strip())
                        st.session_state[f"renaming_{doc_id}"] = False
                        if st.session_state.current_doc_id == doc_id:
                            pass  # refrescará solo
                        st.rerun()
                if rn2.form_submit_button("✕", use_container_width=True):
                    st.session_state[f"renaming_{doc_id}"] = False
                    st.rerun()

        # ── Cambiar estado ──
        status_map = {s: s for s in db.DOCUMENT_STATUSES}
        new_sts = st.selectbox(
            "Estado",
            db.DOCUMENT_STATUSES,
            index=_safe_idx(db.DOCUMENT_STATUSES, doc.get("status","Por revisar")),
            key=f"sts_sel_{doc_id}",
            label_visibility="collapsed",
        )
        if st.button("✔ Aplicar estado", key=f"apply_sts_{doc_id}", use_container_width=True):
            db.update_document(doc_id, status=new_sts)
            st.rerun()

        # ── Asignar / mover carpetas ──
        all_labels   = ["Sin carpeta"] + [f"{f['icon']} {f['name']}" for f in all_folders]
        curr_folders = db.get_folders_for_document(doc_id)
        curr_ids     = {f["id"] for f in curr_folders}

        selected_labels = st.multiselect(
            "📁 Carpetas",
            options    = [f"{f['icon']} {f['name']}" for f in all_folders],
            default    = [f"{f['icon']} {f['name']}" for f in curr_folders],
            key        = f"folders_ms_{doc_id}",
            label_visibility="collapsed",
        )
        if st.button("✔ Aplicar carpetas", key=f"apply_folders_{doc_id}", use_container_width=True):
            sel_ids = set()
            for lbl in selected_labels:
                for f in all_folders:
                    if f"{f['icon']} {f['name']}" == lbl:
                        sel_ids.add(f["id"])
            for fid in sel_ids - curr_ids:
                db.assign_document_to_folder(doc_id, fid)
            for fid in curr_ids - sel_ids:
                db.remove_document_from_folder(doc_id, fid)
            st.session_state[f"doc_menu_{doc_id}"] = False
            st.rerun()

        # ── Eliminar ──
        if st.button("🗑 Quitar de biblioteca", key=f"del_doc_{doc_id}", use_container_width=True):
            db.delete_document(doc_id)
            if st.session_state.current_doc_id == doc_id:
                st.session_state.current_doc_id = None
            st.rerun()


def _render_search_results(query: str):
    results = db.search_documents_fts(query)
    # Fallback: nombre
    if not results:
        all_docs = db.get_all_documents()
        results  = [
            {"doc_id": str(d["id"]), "snippet": d.get("path","")}
            for d in all_docs
            if query.lower() in d.get("name","").lower()
        ]

    st.sidebar.markdown(f"**Resultados para: _{query}_**")
    if not results:
        st.sidebar.info("Sin resultados.")
        return

    for r in results:
        doc_id = int(r["doc_id"])
        doc    = db.get_document(doc_id)
        if not doc:
            continue
        fmt  = doc.get("format","txt").lower()
        icon = FMT_ICONS.get(fmt,"📄")
        if st.sidebar.button(f"{icon} {doc['name'][:30]}", key=f"sr_{doc_id}",
                              use_container_width=True):
            st.session_state.current_doc_id   = doc_id
            st.session_state.show_add_comment = False
            db.update_last_opened(doc_id)
            st.rerun()


# ─────────────────────────────────────────────────────────────────────────────
# Exportación
# ─────────────────────────────────────────────────────────────────────────────

def _export_excel_bytes(comments) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Comentarios"
        headers = ["Documento","Autor","Categoría","Prioridad","Estado",
                   "Página","Apartado comentado","Comentario","Fecha"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1e3a5f")
            cell.alignment = Alignment(horizontal="center")
        for row, c in enumerate(comments, 2):
            ws.append([
                c.get("doc_name",""),
                c.get("author",""),
                c.get("category",""),
                c.get("priority",""),
                c.get("status",""),
                c.get("location_info",""),
                c.get("highlighted_text",""),
                c.get("content",""),
                (c.get("created_at","") or "")[:16],
            ])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except ImportError:
        return b""


def _export_csv_bytes(comments) -> bytes:
    import csv
    buf = io.StringIO()
    w   = csv.writer(buf)
    w.writerow(["Documento","Autor","Categoría","Prioridad","Estado",
                "Página","Apartado comentado","Comentario","Fecha"])
    for c in comments:
        w.writerow([
            c.get("doc_name",""), c.get("author",""), c.get("category",""),
            c.get("priority",""), c.get("status",""), c.get("location_info",""),
            c.get("highlighted_text",""), c.get("content",""),
            (c.get("created_at","") or "")[:16],
        ])
    return buf.getvalue().encode("utf-8-sig")


def render_export_section():
    st.markdown("---")
    with st.expander("📤 Exportar comentarios", expanded=False):
        all_docs   = db.get_all_documents()
        doc_labels = ["Todos los documentos"] + [d.get("name","") for d in all_docs]
        doc_choice = st.selectbox("Documento", doc_labels, key="exp_doc")
        fmt_choice = st.radio("Formato", ["Excel (.xlsx)","CSV (.csv)"], horizontal=True, key="exp_fmt")

        doc_id = None
        if doc_choice != "Todos los documentos":
            idx    = doc_labels.index(doc_choice) - 1
            doc_id = all_docs[idx]["id"]

        comments = db.get_all_comments_for_export(doc_id=doc_id)
        if not comments:
            st.info("No hay comentarios para exportar.")
        else:
            fname = f"comentarios_{datetime.now().strftime('%Y%m%d_%H%M')}"
            if fmt_choice.startswith("Excel"):
                data = _export_excel_bytes(comments)
                if data:
                    st.download_button("⬇️ Descargar Excel", data=data,
                                       file_name=f"{fname}.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                       key="dl_excel")
                else:
                    st.warning("Instala openpyxl: `pip install openpyxl`")
            else:
                data = _export_csv_bytes(comments)
                st.download_button("⬇️ Descargar CSV", data=data,
                                   file_name=f"{fname}.csv",
                                   mime="text/csv",
                                   key="dl_csv")


# ─────────────────────────────────────────────────────────────────────────────
# Pantalla de bienvenida
# ─────────────────────────────────────────────────────────────────────────────

def render_welcome():
    st.markdown("""
    <div style="text-align:center;padding:80px 20px;color:#cdd6f4;">
        <div style="font-size:72px;">📚</div>
        <h1 style="font-size:38px;font-weight:700;margin:16px 0 8px 0;color:#89b4fa;">
            DocManager
        </h1>
        <p style="font-size:17px;color:#6c7086;max-width:480px;margin:0 auto;">
            Tu espacio para leer, anotar y organizar documentos.<br>
            Añade documentos desde el panel izquierdo para comenzar.
        </p>
        <div style="margin-top:40px;display:flex;gap:16px;justify-content:center;flex-wrap:wrap;">
            <div style="background:#24273a;border:1px solid #313244;border-radius:10px;
                        padding:20px 28px;min-width:160px;">
                <div style="font-size:28px;">📁</div>
                <div style="color:#89b4fa;font-weight:600;margin-top:8px;">Carpetas anidadas</div>
                <div style="color:#6c7086;font-size:12px;margin-top:4px;">Subcarpetas ilimitadas</div>
            </div>
            <div style="background:#24273a;border:1px solid #313244;border-radius:10px;
                        padding:20px 28px;min-width:160px;">
                <div style="font-size:28px;">💬</div>
                <div style="color:#89b4fa;font-weight:600;margin-top:8px;">Comentarios</div>
                <div style="color:#6c7086;font-size:12px;margin-top:4px;">Con apartado comentado</div>
            </div>
            <div style="background:#24273a;border:1px solid #313244;border-radius:10px;
                        padding:20px 28px;min-width:160px;">
                <div style="font-size:28px;">🗂️</div>
                <div style="color:#89b4fa;font-weight:600;margin-top:8px;">Repositorio</div>
                <div style="color:#6c7086;font-size:12px;margin-top:4px;">Organizado por carpeta</div>
            </div>
            <div style="background:#24273a;border:1px solid #313244;border-radius:10px;
                        padding:20px 28px;min-width:160px;">
                <div style="font-size:28px;">📤</div>
                <div style="color:#89b4fa;font-weight:600;margin-top:8px;">Exportar</div>
                <div style="color:#6c7086;font-size:12px;margin-top:4px;">Excel y CSV</div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# Utilidades
# ─────────────────────────────────────────────────────────────────────────────

def _safe_idx(lst, val):
    try:
        return lst.index(val)
    except ValueError:
        return 0


def _extract_page(loc: str) -> str:
    if not loc:
        return ""
    m = re.search(r'[Pp]ágina\s*(\d+)', loc)
    if m:
        return m.group(1)
    m = re.search(r'(\d+)', loc)
    return m.group(1) if m else ""


def _folder_depth(all_folders, folder_id, depth=0):
    f = next((x for x in all_folders if x["id"] == folder_id), None)
    if not f or f.get("parent_id") is None:
        return depth
    return _folder_depth(all_folders, f["parent_id"], depth + 1)


# ─────────────────────────────────────────────────────────────────────────────
# Punto de entrada
# ─────────────────────────────────────────────────────────────────────────────

def main():
    _init_state()

    # Inicializar BD
    db.init_db()
    db.init_folders()
    db.add_author_column()
    db.add_parent_id_column()
    db.migrate_existing_to_vault()
    db.auto_create_folders_from_tags()

    # ── Sidebar ──────────────────────────────────────────────────────────────
    with st.sidebar:
        render_sidebar()
        render_export_section()

    # ── Área principal ───────────────────────────────────────────────────────
    doc_id = st.session_state.current_doc_id
    if doc_id:
        doc = db.get_document(doc_id)
        if doc:
            # Encabezado del documento
            st.markdown(
                f'<div style="display:flex;align-items:center;gap:10px;margin-bottom:8px;">'
                f'<span style="font-size:22px;">'
                f'{FMT_ICONS.get(doc.get("format","txt").lower(),"📄")}</span>'
                f'<span style="font-size:20px;font-weight:700;color:#cdd6f4;">{doc["name"]}</span>'
                f'<span style="background:{DOC_STATUS_COLORS.get(doc.get("status","Por revisar"),"#89b4fa")}22;'
                f'color:{DOC_STATUS_COLORS.get(doc.get("status","Por revisar"),"#89b4fa")};'
                f'border-radius:6px;padding:2px 8px;font-size:12px;">{doc.get("status","")}</span>'
                f'</div>',
                unsafe_allow_html=True,
            )

            # ── Barra de herramientas del documento ─────────────────────
            tb_left, tb_right = st.columns([8, 1])
            with tb_right:
                cmt_visible = st.session_state.get("comments_visible", True)
                cmt_icon    = "💬" if cmt_visible else "💬 +"
                if st.button(cmt_icon, key="toggle_comments",
                             help="Mostrar / ocultar comentarios",
                             use_container_width=True):
                    st.session_state.comments_visible = not cmt_visible
                    st.rerun()

            # ── Layout dinámico según visibilidad de comentarios ─────────────
            if st.session_state.get("comments_visible", True):
                col_v, col_c = st.columns([3, 2])
                with col_v:
                    render_viewer(doc)
                with col_c:
                    render_comments(doc)
            else:
                render_viewer(doc)
        else:
            st.session_state.current_doc_id = None
            render_welcome()
    else:
        render_welcome()


if __name__ == "__main__":
    main()
