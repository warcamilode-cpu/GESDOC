# -*- coding: utf-8 -*-
"""utils/exporter.py — Export comments to Excel and PDF"""
import os
import re
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                     Paragraph, Spacer, HRFlowable)
    from reportlab.lib.enums import TA_CENTER
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False


def _get_page(comment):
    """Extract page number string from a comment dict.
    
    Accepts 'Pagina 5', 'Página 5', 'Pag 5', 'pag. 3', etc.
    Returns the number as a string, or empty string if not found.
    """
    loc = comment.get("location_info") or ""
    if not loc:
        return ""
    # Match: Pagina 5 / Página 5 / Pag 5 / pag. 3 (ASCII and latin-1 á)
    m = re.search(r'[Pp]ag(?:ina)?[\s.:]*(\d+)', loc, re.IGNORECASE)
    if m:
        return m.group(1)
    # Fallback: bare number at start
    m = re.search(r'^(\d+)', loc.strip())
    return m.group(1) if m else ""


# ─────────────────────────────────────────────────────────────────────────────
# Excel export
# ─────────────────────────────────────────────────────────────────────────────

def export_to_excel(comments, output_path, title="Comentarios"):
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl no está instalado. Ejecuta: pip install openpyxl")

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Comentarios"

    # Styles
    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill = PatternFill("solid", fgColor="EEF2F8")
    thin     = Side(style="thin", color="CCCCCC")
    bdr      = Border(left=thin, right=thin, top=thin, bottom=thin)
    pri_clrs = {"Alta": "F38BA8", "Media": "F9E2AF", "Baja": "A6E3A1"}
    sts_clrs = {"Abierto": "89B4FA", "Resuelto": "A6E3A1", "Pendiente": "FAB387"}

    # Column definitions: (header, field_fn, width)
    COLS = [
        ("#",             lambda i, c: i,                                    4),
        ("Documento",     lambda i, c: c.get("doc_name", "—"),              22),
        ("Categoría",     lambda i, c: c.get("category", "—"),              12),
        ("Prioridad",     lambda i, c: c.get("priority", "—"),              10),
        ("Estado",        lambda i, c: c.get("status",   "—"),              10),
        ("Página",        lambda i, c: _get_page(c) or "—",                  7),
        ("Autor",         lambda i, c: c.get("author") or "—",              16),
        ("Texto Resaltado", lambda i, c: c.get("highlighted_text") or "",   30),
        ("Comentario",    lambda i, c: c.get("content") or "",              44),
        ("Fecha",         lambda i, c: (c.get("created_at") or "")[:16],   16),
    ]
    NCOLS = len(COLS)

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(NCOLS)}1")
    tc = ws["A1"]
    tc.value     = f"Comentarios exportados — {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    tc.font      = Font(bold=True, size=13, color="1E3A5F")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Header row
    for col, (hdr, _, _w) in enumerate(COLS, 1):
        cell           = ws.cell(row=2, column=col, value=hdr)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.border    = bdr
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
    ws.row_dimensions[2].height = 22

    # Data rows
    for i, comment in enumerate(comments, 1):
        row = i + 2
        for col, (_, fn, _w) in enumerate(COLS, 1):
            val            = fn(i, comment)
            cell           = ws.cell(row=row, column=col, value=val)
            cell.border    = bdr
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if i % 2 == 0:
                cell.fill = alt_fill

        # Priority colour
        pri = comment.get("priority", "")
        if pri in pri_clrs:
            ws.cell(row=row, column=4).fill = PatternFill("solid", fgColor=pri_clrs[pri])
        # Status colour
        sts = comment.get("status", "")
        if sts in sts_clrs:
            ws.cell(row=row, column=5).fill = PatternFill("solid", fgColor=sts_clrs[sts])

        ws.row_dimensions[row].height = 42

    # Column widths
    for col, (_, _, w) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Summary sheet
    from collections import Counter
    ws2 = wb.create_sheet("Resumen")
    ws2["A1"] = "Resumen de comentarios"
    ws2["A1"].font = Font(bold=True, size=13, color="1E3A5F")

    def summary_block(start_row, heading, counter):
        ws2.cell(row=start_row, column=1, value=heading).font = Font(bold=True)
        for offset, (k, v) in enumerate(sorted(counter.items()), 1):
            ws2.cell(row=start_row + offset, column=1, value=f"  {k or '—'}")
            ws2.cell(row=start_row + offset, column=2, value=v)
        return start_row + len(counter) + 2

    r = summary_block(3,  "Total", Counter({"Comentarios": len(comments)}))
    r = summary_block(r,  "Por Prioridad", Counter(c.get("priority") for c in comments))
    r = summary_block(r,  "Por Estado",    Counter(c.get("status")   for c in comments))
    r = summary_block(r,  "Por Autor",     Counter(c.get("author") or "Sin autor"
                                                   for c in comments))

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 14

    wb.save(output_path)
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
# PDF export
# ─────────────────────────────────────────────────────────────────────────────

def export_to_pdf(comments, output_path, title="Comentarios"):
    if not HAS_REPORTLAB:
        raise RuntimeError("reportlab no está instalado. Ejecuta: pip install reportlab")

    doc    = SimpleDocTemplate(output_path, pagesize=A4,
                               rightMargin=1.5*cm, leftMargin=1.5*cm,
                               topMargin=2*cm,     bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story  = []

    # Title
    story.append(Paragraph(
        f"Comentarios exportados",
        ParagraphStyle("T", parent=styles["Title"], fontSize=16,
                       textColor=colors.HexColor("#1e3a5f"), alignment=TA_CENTER,
                       spaceAfter=4)
    ))
    story.append(Paragraph(
        f"Exportado el {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        ParagraphStyle("S", parent=styles["Normal"], fontSize=9,
                       textColor=colors.gray, alignment=TA_CENTER, spaceAfter=10)
    ))
    story.append(HRFlowable(width="100%", thickness=1,
                             color=colors.HexColor("#89b4fa")))
    story.append(Spacer(1, 10))

    if not comments:
        story.append(Paragraph("No hay comentarios.", styles["Normal"]))
        doc.build(story)
        return output_path

    cs   = ParagraphStyle("C", parent=styles["Normal"], fontSize=7, leading=9)
    rows = [["#", "Documento", "Cat.", "Pri.", "Est.", "Pág.", "Autor",
             "Comentario", "Fecha"]]

    pri_pdf = {"Alta": colors.HexColor("#f38ba8"),
               "Media": colors.HexColor("#f9e2af"),
               "Baja":  colors.HexColor("#a6e3a1")}

    for i, c in enumerate(comments, 1):
        rows.append([
            str(i),
            Paragraph(c.get("doc_name", "—"), cs),
            c.get("category", "—"),
            c.get("priority", "—"),
            c.get("status",   "—"),
            _get_page(c) or "—",
            Paragraph(c.get("author") or "—", cs),
            Paragraph((c.get("content") or "")[:200], cs),
            (c.get("created_at") or "")[:10],
        ])

    widths = [0.6*cm, 3.0*cm, 1.8*cm, 1.5*cm, 1.8*cm,
              0.8*cm, 2.5*cm, 4.8*cm, 1.8*cm]
    t = Table(rows, colWidths=widths, repeatRows=1)
    tstyle = [
        ("BACKGROUND", (0,0),  (-1,0), colors.HexColor("#1e3a5f")),
        ("TEXTCOLOR",  (0,0),  (-1,0), colors.white),
        ("FONTNAME",   (0,0),  (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0),  (-1,-1), 7),
        ("ALIGN",      (0,0),  (-1,-1), "LEFT"),
        ("VALIGN",     (0,0),  (-1,-1), "TOP"),
        ("GRID",       (0,0),  (-1,-1), 0.4, colors.HexColor("#cccccc")),
        ("ROWBACKGROUNDS", (0,1), (-1,-1),
         [colors.white, colors.HexColor("#eef2f8")]),
        ("PADDING",    (0,0),  (-1,-1), 4),
    ]
    for i, c in enumerate(comments, 1):
        pc = pri_pdf.get(c.get("priority",""))
        if pc:
            tstyle.append(("BACKGROUND", (3,i), (3,i), pc))
    t.setStyle(TableStyle(tstyle))
    story.append(t)

    doc.build(story)
    return output_path
